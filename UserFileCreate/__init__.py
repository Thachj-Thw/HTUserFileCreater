from dataclasses import dataclass, field
from openpyxl import Workbook, load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.styles import borders
from openpyxl.cell import Cell
import module


@dataclass
class Transportation:
    types: str
    coloumn: str


@dataclass
class InputStruct:
    start_row: int = 5
    name_column: str = "B"
    room_column: str = "B"
    car: Transportation = field(default_factory=lambda: Transportation("Ô tô", "C"))
    moto: Transportation = field(default_factory=lambda: Transportation("Xe máy", "D"))
    e_moto: Transportation = field(default_factory=lambda: Transportation("Xe máy điện", "E"))


@dataclass
class OutputStruct:
    start_row: int = 5
    ordinal_column: str = "A"
    name_column: str = "B"
    room_column: str = "C"
    position_column: str = "D"
    number_column: str = "E"
    transportation_column: str = "P"
    card_column: str = "X"
    end_column: str = "AB"


@dataclass
class RoomInformation:
    row: int
    name: str


@dataclass
class LicensePlateInfomation:
    user: str
    room: str
    number: str
    type: str
    position: str


class UserFileCreater:
    path = module.Path(__file__)


    def __init__(self, file_path: str, input_struct: InputStruct = None, output_struct: OutputStruct = None) -> None:
        self.workbook = load_workbook(filename=file_path)
        self.sheet = self.workbook.active
        self.input_struct = input_struct if input_struct else InputStruct()
        self.output_struct = output_struct if output_struct else OutputStruct()
        self.save_row = self.output_struct.start_row
        self.save_file = load_workbook(self.path.source.join('BaseOutput.xlsx'))
        self.save_sheet = self.save_file.active
        self.list_license = []
    
    def save(self, save_as: str = "output.xlsx") -> None:
        self.save_file.save(filename=save_as)
    
    def detect_room(self) -> list[RoomInformation]:
        rooms = []
        for cell in self.sheet[self.input_struct.room_column]:
            if cell.row < self.input_struct.start_row:
                continue
            fgColor = cell.fill.fgColor.rgb
            if type(fgColor) is str and fgColor != COLOR_INDEX[0]:
                rooms.append(RoomInformation(cell.row, cell.value))
        return rooms

    def create(self) -> None:
        rooms = self.detect_room()
        idx = 0
        max_idx = len(rooms) - 1
        for cell in self.sheet[self.input_struct.name_column]:
            if cell.row < self.input_struct.start_row:
                continue
            if idx < max_idx and cell.row == rooms[idx + 1].row:
                idx += 1
                continue
            self._split_license_plate(cell, rooms[idx].name, self.input_struct.car)
            self._split_license_plate(cell, rooms[idx].name, self.input_struct.moto)
            self._split_license_plate(cell, rooms[idx].name, self.input_struct.e_moto)
    
    def _split_license_plate(self, cell: Cell, room: str, transportation: Transportation) -> None:
        if license_plates := self.sheet[transportation.coloumn + str(cell.row)].value:
            for license_plate in license_plates.split("\n"):
                if not license_plate:
                    continue
                infomation = LicensePlateInfomation(cell.value, room, license_plate, transportation.types, "Nhân viên")
                self.list_license.append(infomation)
                self._set_save_cell(self.output_struct.ordinal_column + str(self.save_row), 1 + self.save_row - self.output_struct.start_row)
                self._set_save_cell(self.output_struct.name_column + str(self.save_row), infomation.user, alignment="left")
                self._set_save_cell(self.output_struct.room_column + str(self.save_row), infomation.room)
                self._set_save_cell(self.output_struct.number_column + str(self.save_row), infomation.number)
                self._set_save_cell(self.output_struct.transportation_column + str(self.save_row), infomation.type)
                self._set_save_cell(self.output_struct.position_column + str(self.save_row), infomation.position)
                self.save_row += 1

    def _set_save_cell(self, dimenstion: str, value: str = "", bold: bool = False, font_color: str = "00000000", size: int = 12, alignment: str = "center"):
        font = Font(bold=bold, color=font_color, size=size)
        alig = Alignment(horizontal=alignment)
        border_type = Side(border_style=borders.BORDER_THIN)
        bor = Border(top=border_type, right=border_type, bottom=border_type, left=border_type)
        cell = self.save_sheet[dimenstion]
        cell.font = font
        cell.alignment = alig
        cell.border = bor
        cell.value = value
